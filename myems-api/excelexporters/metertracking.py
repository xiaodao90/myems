import base64
import uuid
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import Workbook


########################################################################################################################
# PROCEDURES
# Step 1: Validate the report data
# Step 2: Generate excelexporters file
# Step 3: Encode the excelexporters file to Base64
########################################################################################################################

def export(result, space_name, reporting_start_datetime_local, reporting_end_datetime_local):
    ####################################################################################################################
    # Step 1: Validate the report data
    ####################################################################################################################
    if result is None:
        return None

    ####################################################################################################################
    # Step 2: Generate excel file from the report data
    ####################################################################################################################
    filename = generate_excel(result,
                              space_name,
                              reporting_start_datetime_local,
                              reporting_end_datetime_local)
    ####################################################################################################################
    # Step 3: Encode the excel file to Base64
    ####################################################################################################################
    binary_file_data = b''
    try:
        with open(filename, 'rb') as binary_file:
            binary_file_data = binary_file.read()
    except IOError as ex:
        pass

    # Base64 encode the bytes
    base64_encoded_data = base64.b64encode(binary_file_data)
    # get the Base64 encoded data using human-readable characters.
    base64_message = base64_encoded_data.decode('utf-8')
    # delete the file from server
    try:
        os.remove(filename)
    except NotImplementedError as ex:
        pass
    return base64_message


def generate_excel(report, space_name, reporting_start_datetime_local, reporting_end_datetime_local):

    wb = Workbook()
    ws = wb.active
    ws.title = "MeterTracking"

    # Row height
    ws.row_dimensions[1].height = 102
    for i in range(2, 5 + 1):
        ws.row_dimensions[i].height = 42

    for i in range(6, len(report['meters']) + 15):
        ws.row_dimensions[i].height = 60

    # Col width
    ws.column_dimensions['A'].width = 1.5

    ws.column_dimensions['B'].width = 25.0

    for i in range(ord('C'), ord('L')):
        ws.column_dimensions[chr(i)].width = 15.0

    # Font
    name_font = Font(name='Arial', size=15, bold=True)
    title_font = Font(name='Arial', size=15, bold=True)

    table_fill = PatternFill(fill_type='solid', fgColor='1F497D')
    f_border = Border(left=Side(border_style='medium', color='00000000'),
                      right=Side(border_style='medium', color='00000000'),
                      bottom=Side(border_style='medium', color='00000000'),
                      top=Side(border_style='medium', color='00000000')
                      )
    b_border = Border(
        bottom=Side(border_style='medium', color='00000000'),
    )

    b_c_alignment = Alignment(vertical='bottom',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    c_c_alignment = Alignment(vertical='center',
                              horizontal='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
    b_r_alignment = Alignment(vertical='bottom',
                              horizontal='right',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)

    # Img
    img = Image("excelexporters/myems.png")
    ws.add_image(img, 'A1')

    # Title
    ws['B3'].alignment = b_r_alignment
    ws['B3'] = 'Space:'
    ws['C3'].border = b_border
    ws['C3'].alignment = b_c_alignment
    ws['C3'] = space_name

    ws['B4'].alignment = b_r_alignment
    ws['B4'] = 'Reporting Start Datetime:'
    ws['C4'].border = b_border
    ws['C4'].alignment = b_c_alignment
    ws['C4'] = reporting_start_datetime_local

    ws['B5'].alignment = b_r_alignment
    ws['B5'] = 'Reporting End Datetime:'
    ws['C5'].border = b_border
    ws['C5'].alignment = b_c_alignment
    ws['C5'] = reporting_end_datetime_local

    # Title
    ws['B6'].border = f_border
    ws['B6'].font = name_font
    ws['B6'].alignment = c_c_alignment
    ws['B6'].fill = table_fill
    ws['B6'] = 'Name'

    ws['C6'].border = f_border
    ws['C6'].alignment = c_c_alignment
    ws['C6'].font = name_font
    ws['C6'].fill = table_fill
    ws['C6'] = 'Space'

    ws['D6'].border = f_border
    ws['D6'].font = name_font
    ws['D6'].alignment = c_c_alignment
    ws['D6'].fill = table_fill
    ws['D6'] = 'Cost Center'

    ws['E6'].border = f_border
    ws['E6'].alignment = c_c_alignment
    ws['E6'].font = name_font
    ws['E6'].fill = table_fill
    ws['E6'] = 'Energy Category'

    ws['F6'].border = f_border
    ws['F6'].font = name_font
    ws['F6'].alignment = c_c_alignment
    ws['F6'].fill = table_fill
    ws['F6'] = 'Description'

    ws['G6'].border = f_border
    ws['G6'].font = name_font
    ws['G6'].alignment = c_c_alignment
    ws['G6'].fill = table_fill
    ws['G6'] = 'Start Value'

    ws['H6'].border = f_border
    ws['H6'].font = name_font
    ws['H6'].alignment = c_c_alignment
    ws['H6'].fill = table_fill
    ws['H6'] = 'End Value'

    current_row_number = 7
    for i in range(0, len(report['meters'])):

        ws['B' + str(current_row_number)].font = title_font
        ws['B' + str(current_row_number)].border = f_border
        ws['B' + str(current_row_number)].alignment = c_c_alignment
        ws['B' + str(current_row_number)] = report['meters'][i]['meter_name']

        ws['C' + str(current_row_number)].font = title_font
        ws['C' + str(current_row_number)].border = f_border
        ws['C' + str(current_row_number)].alignment = c_c_alignment
        ws['C' + str(current_row_number)] = report['meters'][i]['space_name']

        ws['D' + str(current_row_number)].font = title_font
        ws['D' + str(current_row_number)].border = f_border
        ws['D' + str(current_row_number)].alignment = c_c_alignment
        ws['D' + str(current_row_number)] = report['meters'][i]['cost_center_name']

        ws['E' + str(current_row_number)].font = title_font
        ws['E' + str(current_row_number)].border = f_border
        ws['E' + str(current_row_number)].alignment = c_c_alignment
        ws['E' + str(current_row_number)] = report['meters'][i]['energy_category_name']

        ws['F' + str(current_row_number)].font = title_font
        ws['F' + str(current_row_number)].border = f_border
        ws['F' + str(current_row_number)].alignment = c_c_alignment
        ws['F' + str(current_row_number)] = report['meters'][i]['description']

        ws['G' + str(current_row_number)].font = title_font
        ws['G' + str(current_row_number)].border = f_border
        ws['G' + str(current_row_number)].alignment = c_c_alignment
        ws['G' + str(current_row_number)] = report['meters'][i]['start_value']

        ws['H' + str(current_row_number)].font = title_font
        ws['H' + str(current_row_number)].border = f_border
        ws['H' + str(current_row_number)].alignment = c_c_alignment
        ws['H' + str(current_row_number)] = report['meters'][i]['end_value']

        current_row_number += 1

    filename = str(uuid.uuid4()) + '.xlsx'
    wb.save(filename)

    return filename
